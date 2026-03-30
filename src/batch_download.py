"""
MOPS PDF Packager — 批次下載模組

讀取「公司清單.xlsx」中的股票代碼與年份區間，
依序下載每家公司的所有 MOPS 報告（年報、財報、三書表、法說會簡報、ESG 永續報告書、提升企業價值計劃）。

功能：
  - 自動偵測已下載的公司資料夾（有 PDF 就跳過，不需手動管理進度檔）
  - 每家公司之間隨機等待 45~90 秒，避免被 MOPS 封鎖
  - 所有 log 同時輸出到 console 和 log 檔案
  - 啟動時自動清除舊 log，只保留本次執行的 log
  - 跑完後印出 summary（成功/失敗統計）
  - 支援自訂儲存路徑（Excel B1 欄位），可指向 Google Drive 資料夾
"""

import sys
import os
import glob
import time
import random
import datetime
import logging
import subprocess

# 將專案根目錄加入路徑（src/ 的上一層）
ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT_DIR)

from src.utils.downloader import MOPSDownloader


# ============================================================
# 設定
# ============================================================

EXCEL_FILENAME = "公司清單.xlsx"
OUTPUT_DIR_NAME = "MOPS_批次下載"


# ============================================================
# Logging 設定
# ============================================================

def cleanup_old_logs(output_dir):
    """刪除輸出目錄中所有舊的 .log 檔案。"""
    if not os.path.exists(output_dir):
        return
    for log_file in glob.glob(os.path.join(output_dir, "*.log")):
        try:
            os.remove(log_file)
        except OSError:
            pass


def setup_logging(output_dir):
    """
    設定 logging：同時輸出到 console 和 log 檔案。
    Log 檔案存放在 output_dir 下。
    """
    os.makedirs(output_dir, exist_ok=True)

    # 清除舊 log
    cleanup_old_logs(output_dir)

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    log_filename = f"batch_download_{timestamp}.log"
    log_path = os.path.join(output_dir, log_filename)

    # 建立 root logger
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    # 清除既有 handler（避免重複）
    logger.handlers.clear()

    formatter = logging.Formatter(
        "[%(asctime)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )

    # Console handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

    # File handler
    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    # 將 print 的輸出也 capture 到 log 檔案
    # 透過覆寫 sys.stdout 讓現有 scraper 的 print() 也能寫入 log
    sys.stdout = _PrintLogger(logger, log_path)

    logging.info(f"Log 檔案: {log_path}")
    return logger


class _PrintLogger:
    """
    攔截 print() 輸出，同時寫入 console 和 log 檔案。
    讓現有 scraper 裡的 print() 也能被 capture。
    """
    def __init__(self, logger, log_path):
        self._logger = logger
        self._log_file = open(log_path, "a", encoding="utf-8")
        self._original_stdout = sys.__stdout__

    def write(self, message):
        if message and message.strip():
            self._original_stdout.write(message)
            if not message.endswith("\n"):
                self._original_stdout.write("\n")
            self._log_file.write(message)
            if not message.endswith("\n"):
                self._log_file.write("\n")
            self._log_file.flush()
        elif message == "\n":
            self._original_stdout.write(message)
            self._log_file.write(message)
            self._log_file.flush()

    def flush(self):
        self._original_stdout.flush()
        self._log_file.flush()

    def close(self):
        self._log_file.close()


# ============================================================
# Excel 讀取
# ============================================================

ALL_REPORT_TYPES = ["年報", "財報", "關係企業三書表", "法說會簡報", "ESG永續報告書", "提升企業價值計劃"]

# Excel D~J 欄對應的報告類型（索引 3~9）
_REPORT_COL_MAP = {
    3: "全選",
    4: "年報",
    5: "財報",
    6: "關係企業三書表",
    7: "法說會簡報",
    8: "ESG永續報告書",
    9: "提升企業價值計劃",
}


def read_company_list(excel_path):
    """
    讀取公司清單 Excel，回傳 (custom_output_dir, companies)。

    Excel 格式：
      第 1 列: A1="儲存路徑（選填）", B1=自訂儲存路徑（留空則使用預設）
      第 2 列: 標題列（股票代碼 / 起始年份 / 結束年份 / 全選 / ...）
      第 3 列起: 資料列
        A 欄: 股票代碼 (必填)
        B 欄: 起始年份 (民國年, 選填)
        C 欄: 結束年份 (民國年, 選填)
        D 欄: 全選 (✓ = 下載全部報告類型)
        E~J 欄: 年報 / 財報 / 關係企業三書表 / 法說會簡報 / ESG永續報告書 / 提升企業價值計劃 (✓ = 下載)

    Returns:
        tuple: (custom_output_dir: str or None, companies: list[dict])
    """
    try:
        import openpyxl
    except ImportError:
        logging.error("缺少 openpyxl 套件。請執行: pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(excel_path):
        logging.error(f"找不到公司清單: {excel_path}")
        logging.error("請先建立「公司清單.xlsx」，格式：A欄=股票代碼, B欄=起始年份(民國年), C欄=結束年份(民國年)")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb.active

    # 自動偵測 Excel 格式：
    #   新格式: Row 1 = 設定列（A1="儲存路徑（選填）"）, Row 2 = 標題, Row 3+ = 資料
    #   舊格式: Row 1 = 標題（A1="股票代碼"）, Row 2+ = 資料
    custom_output_dir = None
    row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    a1_val = str(row1[0][0]).strip() if row1 and row1[0] and row1[0][0] else ""

    if "儲存路徑" in a1_val:
        # 新格式：有設定列
        data_start_row = 3
        if len(row1[0]) > 1 and row1[0][1]:
            path_val = str(row1[0][1]).strip()
            if path_val:
                custom_output_dir = path_val
    else:
        # 舊格式：直接從第 2 列開始讀資料
        data_start_row = 2

    companies = []
    current_roc_year = datetime.datetime.now().year - 1911

    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        if not row or not row[0]:
            continue  # 跳過空列

        ticker = str(row[0]).strip()
        if not ticker:
            continue

        # 讀取起始/結束年份
        year_start = None
        year_end = None

        if len(row) > 1 and row[1] is not None:
            try:
                year_start = int(row[1])
            except (ValueError, TypeError):
                logging.warning(f"第 {row_idx} 列: 起始年份格式錯誤「{row[1]}」，將使用預設值")

        if len(row) > 2 and row[2] is not None:
            try:
                year_end = int(row[2])
            except (ValueError, TypeError):
                logging.warning(f"第 {row_idx} 列: 結束年份格式錯誤「{row[2]}」，將使用預設值")

        # 預設值邏輯
        if year_start is None and year_end is None:
            year_end = current_roc_year
            year_start = current_roc_year - 4
        elif year_start is not None and year_end is None:
            year_end = current_roc_year
        elif year_start is None and year_end is not None:
            year_start = year_end - 4

        # 讀取報告類型 (D~J 欄, 索引 3~9)
        def _is_checked(col_idx):
            if len(row) > col_idx and row[col_idx]:
                val = str(row[col_idx]).strip()
                return val in ("\u2713", "V", "v", "O", "o", "Y", "y", "1", "TRUE", "True", "true")
            return False

        select_all = _is_checked(3)  # D 欄: 全選

        if select_all:
            report_types = list(ALL_REPORT_TYPES)
        else:
            report_types = []
            for col_idx in range(4, 10):  # E~J 欄
                if _is_checked(col_idx):
                    report_types.append(_REPORT_COL_MAP[col_idx])

            # 如果 D~J 全部留空，當作全選（防呆）
            if not report_types:
                report_types = list(ALL_REPORT_TYPES)

        companies.append({
            "ticker": ticker,
            "year_start": year_start,
            "year_end": year_end,
            "report_types": report_types,
        })

    wb.close()
    return custom_output_dir, companies


# ============================================================
# 已下載偵測
# ============================================================

def is_already_downloaded(output_dir, ticker):
    """
    檢查某家公司是否已經下載過。
    判斷標準：output_dir 裡有以 {ticker}_ 開頭的資料夾，且裡面有至少一個 PDF。
    """
    if not os.path.exists(output_dir):
        return False

    for name in os.listdir(output_dir):
        folder_path = os.path.join(output_dir, name)
        if not os.path.isdir(folder_path):
            continue
        # 資料夾名稱格式: {ticker}_{公司名} 或就是 {ticker}
        if name == ticker or name.startswith(f"{ticker}_"):
            # 檢查資料夾內（含子資料夾）有沒有 PDF
            for root, dirs, files in os.walk(folder_path):
                for f in files:
                    if f.lower().endswith(".pdf"):
                        return True
    return False


# ============================================================
# 自動 Push 到 GitHub
# ============================================================

def auto_push_to_github(script_dir):
    """
    下載完成後自動將報告 push 到 GitHub。
    如果 git 未安裝或不是 git repo，會顯示提示並跳過。
    """
    logging.info("\n" + "=" * 60)
    logging.info("自動上傳到 GitHub")
    logging.info("=" * 60)

    # 1. 檢查 git 是否已安裝
    try:
        subprocess.run(
            ["git", "--version"],
            cwd=script_dir, capture_output=True, check=True
        )
    except FileNotFoundError:
        logging.warning("未偵測到 git，跳過自動上傳。")
        logging.warning("如需自動上傳功能，請先安裝 Git: https://git-scm.com/")
        return
    except subprocess.CalledProcessError:
        logging.warning("git 執行異常，跳過自動上傳。")
        return

    # 2. 檢查是否為 git repo
    result = subprocess.run(
        ["git", "rev-parse", "--git-dir"],
        cwd=script_dir, capture_output=True
    )
    if result.returncode != 0:
        logging.warning("此目錄不是 Git 儲存庫，跳過自動上傳。")
        logging.warning("請先執行 git init 並設定 remote。")
        return

    # 3. git add（依照 .gitignore 規則加入所有變更）
    logging.info("正在加入檔案到 Git...")
    subprocess.run(
        ["git", "add", "-A"],
        cwd=script_dir, capture_output=True
    )

    # 4. 檢查是否有變更需要 commit
    status_result = subprocess.run(
        ["git", "diff", "--cached", "--quiet"],
        cwd=script_dir, capture_output=True
    )
    if status_result.returncode == 0:
        logging.info("沒有新的變更需要上傳。")
        return

    # 5. git commit
    today = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    commit_msg = f"更新報告 {today}"
    logging.info(f"正在 commit: {commit_msg}")
    commit_result = subprocess.run(
        ["git", "commit", "-m", commit_msg],
        cwd=script_dir, capture_output=True, text=True
    )
    if commit_result.returncode != 0:
        logging.error(f"git commit 失敗: {commit_result.stderr}")
        return

    # 6. git push
    logging.info("正在 push 到 GitHub...")
    push_result = subprocess.run(
        ["git", "push"],
        cwd=script_dir, capture_output=True, text=True
    )
    if push_result.returncode == 0:
        logging.info("成功上傳到 GitHub！")
    else:
        logging.error(f"git push 失敗: {push_result.stderr}")
        logging.error("請確認網路連線及 GitHub 認證設定。")


# ============================================================
# 主流程
# ============================================================

def main():
    # PyInstaller --onefile 會解壓到暫存目錄，__file__ 指向暫存路徑
    # 必須用 sys.executable 取得 exe 實際所在資料夾
    if getattr(sys, 'frozen', False):
        script_dir = os.path.dirname(sys.executable)
    else:
        script_dir = ROOT_DIR
    excel_path = os.path.join(script_dir, EXCEL_FILENAME)

    # 讀取公司清單（含自訂儲存路徑）
    custom_output_dir, companies = read_company_list(excel_path)

    # 決定輸出目錄
    if custom_output_dir:
        output_dir = custom_output_dir
    else:
        output_dir = os.path.join(script_dir, OUTPUT_DIR_NAME)

    # 設定 logging（會自動清除舊 log）
    logger = setup_logging(output_dir)

    logging.info("=" * 60)
    logging.info("MOPS PDF Packager — 批次下載")
    logging.info("=" * 60)

    if custom_output_dir:
        logging.info(f"自訂儲存路徑: {custom_output_dir}")

    if not companies:
        logging.error("公司清單為空，請確認 Excel 內容。")
        sys.exit(1)

    logging.info(f"共讀取到 {len(companies)} 家公司")
    for c in companies:
        types_str = ", ".join(c["report_types"])
        logging.info(f"  {c['ticker']}  民國 {c['year_start']}~{c['year_end']} 年  [{types_str}]")

    # 開始下載
    success_list = []
    skip_list = []
    fail_list = []

    for idx, company in enumerate(companies, start=1):
        ticker = company["ticker"]

        # 偵測是否已下載過（資料夾內有 PDF）
        if is_already_downloaded(output_dir, ticker):
            logging.info(f"\n[{idx}/{len(companies)}] {ticker} — 已有下載資料，跳過")
            skip_list.append(ticker)
            continue

        logging.info(f"\n{'=' * 60}")
        logging.info(f"[{idx}/{len(companies)}] 開始下載 {ticker} (民國 {company['year_start']}~{company['year_end']} 年)")
        logging.info("=" * 60)

        try:
            downloader = MOPSDownloader(
                ticker=ticker,
                save_base_dir=output_dir,
                year_start=company["year_start"],
                year_end=company["year_end"],
                report_types=company["report_types"],
            )
            downloader.run(use_subdirs=True)

            success_list.append(ticker)
            logging.info(f"[{ticker}] 下載完成！")

        except Exception as e:
            logging.error(f"[{ticker}] 下載過程發生錯誤: {e}")
            fail_list.append(ticker)

        # 如果不是最後一家，等待一段時間
        if idx < len(companies):
            wait_seconds = random.uniform(45, 90)
            logging.info(f"\n等待 {wait_seconds:.0f} 秒後繼續下一家...")
            time.sleep(wait_seconds)

    # ============================================================
    # Summary
    # ============================================================
    logging.info("\n" + "=" * 60)
    logging.info("批次下載完成 — Summary")
    logging.info("=" * 60)
    logging.info(f"本次下載: {len(success_list)} 家")
    logging.info(f"已有資料（跳過）: {len(skip_list)} 家")
    logging.info(f"失敗: {len(fail_list)} 家")

    if skip_list:
        logging.info(f"跳過的公司: {', '.join(skip_list)}")
        logging.info("如需重新下載，請先刪除該公司的資料夾後重跑。")

    if fail_list:
        logging.info(f"失敗的公司: {', '.join(fail_list)}")

    logging.info(f"\n檔案存放於: {os.path.abspath(output_dir)}")

    # 自動上傳到 GitHub
    auto_push_to_github(script_dir)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n發生未預期的錯誤: {e}")
    finally:
        print()
        input("按 Enter 鍵關閉視窗...")
